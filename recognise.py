import cv2
import numpy as np
import pickle
from datetime import datetime, timedelta
import os
from openpyxl import Workbook, load_workbook
import time

# === Configuration ===
ESP32_CAM_URL = "http://192.168.188.140:81/stream"
CONFIDENCE_THRESHOLD = 110  # Lower = stricter. Typical good range: 50–70

month_year = datetime.now().strftime('%B_%Y')
EXCEL_PATH = f"Attendance_{month_year}.xlsx"

# === Load model and label data ===
recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read("trainer/trainer.yml")

face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

with open("trainer/labels.pickle", 'rb') as f:
    label_info = pickle.load(f)  # {id: "name_rollnumber"}


# === Create new Excel sheet if it doesn't exist ===
def create_monthly_sheet():
    if os.path.exists(EXCEL_PATH):
        print(f"📄 Excel file already exists: {EXCEL_PATH} — keeping existing data.")

        # Check and add any missing date columns
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active

        existing_headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]

        start_date = datetime.now().replace(day=1)
        end_date = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

        current_date = start_date
        updated = False
        while current_date <= end_date:
            date_str = current_date.strftime("%d-%m")
            if date_str not in existing_headers:
                ws.cell(row=1, column=ws.max_column + 1).value = date_str
                updated = True
            current_date += timedelta(days=1)

        if updated:
            wb.save(EXCEL_PATH)
            print("🛠 Updated Excel with missing date columns.")
        return

    # Create new sheet if file doesn't exist
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = ["Name", "Roll Number"]
    start_date = datetime.now().replace(day=1)
    end_date = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

    current_date = start_date
    while current_date <= end_date:
        headers.append(current_date.strftime("%d-%m"))
        current_date += timedelta(days=1)

    ws.append(headers)
    wb.save(EXCEL_PATH)
    print(f"📄 Created new attendance file: {EXCEL_PATH}")


# === Update attendance with duplicate check ===
def update_attendance(name, roll_number):
    now = datetime.now()
    date_str = now.strftime('%d-%m')
    time_str = now.strftime('%H:%M:%S')

    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    date_col = None
    for col in range(3, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == date_str:
            date_col = col
            break

    if not date_col:
        print(f"❌ Date column {date_str} not found.")
        return

    student_found = False
    for row in range(2, ws.max_row + 1):
        name_cell = ws.cell(row=row, column=1).value
        roll_cell = ws.cell(row=row, column=2).value
        if name_cell == name and roll_cell == roll_number:
            student_found = True
            if not ws.cell(row=row, column=date_col).value:
                ws.cell(row=row, column=date_col).value = time_str
                print(f"🟢 Marked attendance: {name} ({roll_number}) at {time_str}")
            else:
                print(f"⚠ Already marked: {name} ({roll_number}) on {date_str}")
            break

    if not student_found:
        new_row = [name, roll_number]
        for col in range(3, ws.max_column + 1):
            new_row.append(time_str if col == date_col else "")
        ws.append(new_row)
        print(f"🆕 Added new student: {name} ({roll_number}) at {time_str}")

    wb.save(EXCEL_PATH)


# === Start Program ===
create_monthly_sheet()

print(f"🔍 Connecting to: {ESP32_CAM_URL}")
cap = cv2.VideoCapture(ESP32_CAM_URL)
time.sleep(2)

if not cap.isOpened():
    print("❌ Could not open stream.")
    exit()

print("✅ Stream opened. Press 'q' to exit.")

while True:
    ret, frame = cap.read()
    if not ret:
        print("⚠ Frame not received.")
        continue

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5)

    for (x, y, w, h) in faces:
        roi = gray[y:y+h, x:x+w]
        id_, conf = recognizer.predict(roi)

        if conf < CONFIDENCE_THRESHOLD:
            full_label = label_info.get(id_, None)

            if full_label and "_" in full_label:
                name, roll_number = full_label.split("_", 1)
                update_attendance(name, roll_number)
                label_text = f"{name} ({roll_number})"
                color = (0, 255, 0)
            else:
                label_text = "Unknown"
                color = (0, 0, 255)
        else:
            label_text = "Unknown"
            color = (0, 0, 255)

        # Draw bounding box and label with confidence
        cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
        cv2.putText(frame, f"{label_text} ({int(conf)})", (x, y - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.8, color, 2)

    cv2.imshow("Face Recognition Attendance", frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        print("👋 Exiting program.")
        break

cap.release()
cv2.destroyAllWindows()